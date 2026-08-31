from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Student, Group, Enrollment, Lesson, Attendance, Payment, Book
from users.models import Profile
from decimal import Decimal


def login_view(request):
    """Страница входа"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Неверный логин или пароль')
    
    return render(request, 'dashboard/login.html')


def logout_view(request):
    """Выход"""
    logout(request)
    return redirect('login')


def get_user_role(user):
    """Получаем роль пользователя"""
    try:
        return user.profile.role
    except:
        return 'admin'


@login_required
def dashboard(request):
    """Главный дашборд"""
    role = get_user_role(request.user)
    
    
    if role == 'teacher':
        groups = Group.objects.filter(teacher=request.user, is_active=True)
        total_students = Enrollment.objects.filter(group__teacher=request.user).count()
        today = timezone.localdate()
        today_lessons = Lesson.objects.filter(
            group__teacher=request.user,
            date=today
        ).count()
        
        context = {
            'role': 'teacher',
            'groups': groups,
            'total_students': total_students,
            'today_lessons': today_lessons,
        }
    else:
        groups = Group.objects.filter(is_active=True)
        total_students = Student.objects.count()
        total_groups = groups.count()

        # Топ должников
        top_debtors = []
        students_with_debt = Student.objects.all()
        for student in students_with_debt:
            unpaid = Payment.objects.filter(student=student, is_paid=False)
            total_debt = sum([float(p.amount) - float(p.paid_amount) for p in unpaid])
            if total_debt > 0:
                top_debtors.append({
                    'student': student,
                    'debt': total_debt,
                    'phone': student.phone,
                })
        top_debtors.sort(key=lambda x: x['debt'], reverse=True)
        top_debtors = top_debtors[:5]
        
        # Статистика оплат (текущие циклы)
        today = timezone.localdate()
        current_payments = Payment.objects.filter(start_date__lte=today, end_date__gte=today)
        paid_count = current_payments.filter(is_paid=True).count()
        unpaid_count = current_payments.filter(is_paid=False).count()
        
        # Данные для графика (оплаты по циклам)
        payment_stats = []
        for cycle in range(1, 13):
            count = Payment.objects.filter(cycle_number=cycle, is_paid=True).count()
            payment_stats.append(count)
        
        today_attendance = Attendance.objects.filter(lesson__date=today)
        present_count = today_attendance.filter(status='present').count()
        total_attendance = today_attendance.count()
        
        context = {
            'role': role,
            'groups': groups,
            'total_students': total_students,
            'total_groups': total_groups,
            'paid_count': paid_count,
            'unpaid_count': unpaid_count,
            'payment_stats': payment_stats,
            'present_count': present_count,
            'total_attendance': total_attendance,
            'top_debtors': top_debtors,
        }
    
    return render(request, 'dashboard/dashboard.html', context)


@login_required
def group_detail(request, group_id):
    """Страница группы: ученики, посещаемость, оплаты"""
    group = get_object_or_404(Group, id=group_id)
    role = get_user_role(request.user)
    
    if role == 'teacher' and group.teacher != request.user:
        messages.error(request, 'У вас нет доступа к этой группе')
        return redirect('dashboard')
    
    enrollments = Enrollment.objects.filter(group=group).select_related('student')
    students = [enrollment.student for enrollment in enrollments]
    
    # Получаем или создаём урок на сегодня (только если сегодня день занятий)
    today = timezone.localdate()
    
    day_keywords = {
        0: ['пн', 'понедельник'],
        1: ['вт', 'вторник'],
        2: ['ср', 'сред'],
        3: ['чт', 'четверг'],
        4: ['пт', 'пятниц'],
        5: ['сб', 'суббот'],
        6: ['вс', 'воскрес'],
    }
    
    schedule_text = group.schedule.lower()
    today_weekday = today.weekday()
    
    is_lesson_day = any(keyword in schedule_text for keyword in day_keywords.get(today_weekday, []))
    
    if is_lesson_day:
        lesson, created = Lesson.objects.get_or_create(
            group=group,
            date=today,
            defaults={'topic': ''}
        )
        
        if created:
            for student in students:
                Attendance.objects.get_or_create(
                    lesson=lesson,
                    student=student,
                    defaults={'status': 'absent'}
                )
    else:
        lesson = None
    
    # Получаем посещаемость (только если урок есть)
    attendance_dict = {}
    if lesson:
        attendances = Attendance.objects.filter(lesson=lesson)
        attendance_dict = {att.student_id: att.status for att in attendances}
    
    # Получаем оплаты за текущий цикл
    payments = Payment.objects.filter(
        group=group,
        start_date__lte=today,
        end_date__gte=today
    )
    payment_dict = {pay.student_id: pay.is_paid for pay in payments}
    
    student_data = []
    for student in students:
        enrollment = Enrollment.objects.filter(student=student, group=group).first()
        student_data.append({
            'student': student,
            'attendance_status': attendance_dict.get(student.id, 'absent'),
            'is_paid': payment_dict.get(student.id, False),
            'enrollment': enrollment,
        })
    
    lessons_history = Lesson.objects.filter(group=group).order_by('-date')[:10]
    
    context = {
        'group': group,
        'student_data': student_data,
        'today': today,
        'role': role,
        'lessons_history': lessons_history,
        'lesson': lesson,
        'can_mark_attendance': role in ['teacher', 'developer'],
    }
    
    return render(request, 'dashboard/group_detail.html', context)


@login_required
def mark_attendance(request, group_id):
    """Отметка посещаемости (AJAX)"""
    if request.method == 'POST':
        group = get_object_or_404(Group, id=group_id)
        role = get_user_role(request.user)
        
        if role not in ['teacher', 'developer']:
            return JsonResponse({'success': False, 'error': 'Нет доступа'})
        
        if role == 'teacher' and group.teacher != request.user:
            return JsonResponse({'success': False, 'error': 'Нет доступа'})
        
        today = timezone.localdate()
        day_keywords = {
            0: ['пн', 'понедельник'],
            1: ['вт', 'вторник'],
            2: ['ср', 'сред'],
            3: ['чт', 'четверг'],
            4: ['пт', 'пятниц'],
            5: ['сб', 'суббот'],
            6: ['вс', 'воскрес'],
        }
        schedule_text = group.schedule.lower()
        today_weekday = today.weekday()
        is_lesson_day = any(keyword in schedule_text for keyword in day_keywords.get(today_weekday, []))
        
        if not is_lesson_day:
            return JsonResponse({'success': False, 'error': 'Сегодня нет урока'})
        
        student_id = request.POST.get('student_id')
        status = request.POST.get('status')
        
        lesson, _ = Lesson.objects.get_or_create(
            group=group,
            date=today
        )
        
        attendance, _ = Attendance.objects.get_or_create(
            lesson=lesson,
            student_id=student_id,
            defaults={'status': status}
        )
        attendance.status = status
        attendance.save()
        
        return JsonResponse({'success': True, 'status': status})
    
    return JsonResponse({'success': False})


@login_required
def toggle_payment(request, group_id):
    """Переключение оплаты (AJAX)"""
    if request.method == 'POST':
        group = get_object_or_404(Group, id=group_id)
        role = get_user_role(request.user)
        
        if role == 'teacher':
            return JsonResponse({'success': False, 'error': 'Нет доступа'})
        
        student_id = request.POST.get('student_id')
        today = timezone.localdate()
        
        payment, created = Payment.objects.get_or_create(
            student_id=student_id,
            group=group,
            start_date__lte=today,
            end_date__gte=today,
            defaults={'amount': group.price, 'is_paid': True, 'cycle_number': 1}
        )
        
        if not created:
            payment.is_paid = not payment.is_paid
            payment.save()
        
        return JsonResponse({'success': True, 'is_paid': payment.is_paid})
    
    return JsonResponse({'success': False})


@login_required
def students_list(request):
    """Список всех учеников"""
    role = get_user_role(request.user)

    if role == 'accountant':
        messages.error(request, 'У вас нет доступа')
        return redirect('payment_management')
    
    if role == 'teacher':
        students = Student.objects.filter(
            enrollments__group__teacher=request.user
        ).distinct()
    else:
        students = Student.objects.all()
    
    student_data = []
    for student in students:
        enrollments = Enrollment.objects.filter(student=student).select_related('group')
        groups_list = [e.group.name for e in enrollments]
        student_data.append({
            'student': student,
            'groups': groups_list,
        })
    
    context = {
        'student_data': student_data,
        'role': role,
    }
    
    return render(request, 'dashboard/students_list.html', context)


@login_required
def add_student(request):
    """Добавление ученика"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('students_list')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone', '')
        parent_name = request.POST.get('parent_name', '')
        parent_phone = request.POST.get('parent_phone', '')
        group_ids = request.POST.getlist('groups')
        
        if name:
            student = Student.objects.create(
                name=name,
                phone=phone,
                parent_name=parent_name,
                parent_phone=parent_phone,
                school=request.POST.get('school', ''),
                grade=request.POST.get('grade', ''),
                age=request.POST.get('age') or None,
            )
            
            for group_id in group_ids:
                group = Group.objects.get(id=group_id)
                Enrollment.objects.get_or_create(student=student, group=group)
            
            messages.success(request, f'Ученик {name} добавлен!')
            return redirect('students_list')
    
    groups = Group.objects.filter(is_active=True)
    context = {
        'groups': groups,
    }
    
    return render(request, 'dashboard/add_student.html', context)


@login_required
def payments_list(request):
    """Все оплаты"""
    role = get_user_role(request.user)
    
    payments = Payment.objects.select_related('student', 'group').order_by('-cycle_number')
    
    if role == 'teacher':
        payments = payments.filter(group__teacher=request.user)
    
    context = {
        'payments': payments,
        'role': role,
    }
    
    return render(request, 'dashboard/payments_list.html', context)


@login_required
def lesson_history(request, group_id):
    """История уроков группы"""
    group = get_object_or_404(Group, id=group_id)
    role = get_user_role(request.user)
    
    if role == 'teacher' and group.teacher != request.user:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    lessons = Lesson.objects.filter(group=group).order_by('-date')
    
    lesson_data = []
    for lesson in lessons:
        attendances = Attendance.objects.filter(lesson=lesson).select_related('student')
        
        present_students = []
        absent_students = []
        
        for att in attendances:
            if att.status == 'present':
                present_students.append(att.student.name)
            else:
                absent_students.append(att.student.name)
        
        lesson_data.append({
            'lesson': lesson,
            'present_count': len(present_students),
            'absent_count': len(absent_students),
            'total_count': len(present_students) + len(absent_students),
            'present_students': present_students,
            'absent_students': absent_students,
        })
    
    context = {
        'group': group,
        'lesson_data': lesson_data,
        'role': role,
    }
    
    return render(request, 'dashboard/lesson_history.html', context)


@login_required
def weekly_schedule(request):
    """Расписание на неделю"""
    role = get_user_role(request.user)
    
    # Админ и бухгалтер не видят расписание
    if role in ['admin', 'accountant']:
        messages.error(request, 'Расписание пока недоступно для вашей роли')
        return redirect('dashboard')
    
    if role == 'teacher':
        groups = Group.objects.filter(teacher=request.user, is_active=True)
    else:
        groups = Group.objects.filter(is_active=True)
    
    days = [
        {'name': 'Понедельник', 'short': 'Пн', 'num': 0},
        {'name': 'Вторник', 'short': 'Вт', 'num': 1},
        {'name': 'Среда', 'short': 'Ср', 'num': 2},
        {'name': 'Четверг', 'short': 'Чт', 'num': 3},
        {'name': 'Пятница', 'short': 'Пт', 'num': 4},
        {'name': 'Суббота', 'short': 'Сб', 'num': 5},
        {'name': 'Воскресенье', 'short': 'Вс', 'num': 6},
    ]
    
    today = timezone.localdate()
    monday = today - timedelta(days=today.weekday())
    
    for i, day in enumerate(days):
        day_date = monday + timedelta(days=i)
        day['date'] = day_date.strftime('%d.%m')
        day['full_date'] = day_date.strftime('%d %B')
        day['is_today'] = (day_date == today)
    
    time_slots = []
    for hour in range(6, 23):
        for minute in [0, 30]:
            time_slots.append({
                'hour': hour,
                'minute': minute,
                'label': f'{hour}:{minute:02d}',
            })
    
    import re
    
    schedule_data = []
    for group in groups:
        schedule_text = group.schedule.lower()
        
        day_indexes = []
        day_keywords = {
            0: ['пн', 'понедельник'],
            1: ['вт', 'вторник'],
            2: ['ср', 'сред'],
            3: ['чт', 'четверг'],
            4: ['пт', 'пятниц'],
            5: ['сб', 'суббот'],
            6: ['вс', 'воскрес'],
        }
        
        for day_num, keywords in day_keywords.items():
            if any(keyword in schedule_text for keyword in keywords):
                day_indexes.append(day_num)
        
        if not day_indexes:
            continue
        
        time_match = re.search(r'(\d{1,2})[:.](\d{2})\s*[-–—]\s*(\d{1,2})[:.](\d{2})', schedule_text)
        
        if time_match:
            start_hour = int(time_match.group(1))
            start_minute = int(time_match.group(2))
            end_hour = int(time_match.group(3))
            end_minute = int(time_match.group(4))
        else:
            time_match = re.search(r'(\d{1,2})[:.](\d{2})', schedule_text)
            if time_match:
                start_hour = int(time_match.group(1))
                start_minute = int(time_match.group(2))
                end_hour = start_hour + 1
                end_minute = start_minute + 30
                if end_minute >= 60:
                    end_hour += 1
                    end_minute -= 60
            else:
                start_hour = 18
                start_minute = 0
                end_hour = 19
                end_minute = 30
        
        start_total_minutes = start_hour * 60 + start_minute
        end_total_minutes = end_hour * 60 + end_minute
        duration_slots = (end_total_minutes - start_total_minutes) // 30
        
        if duration_slots < 1:
            duration_slots = 1
        
        for day_index in day_indexes:
            schedule_data.append({
                'group': group,
                'day_index': day_index,
                'start_hour': start_hour,
                'start_minute': start_minute,
                'duration_slots': duration_slots,
            })
    
    pastel_colors = [
        {'bg': 'bg-blue-100', 'border': 'border-blue-300', 'text': 'text-blue-800'},
        {'bg': 'bg-green-100', 'border': 'border-green-300', 'text': 'text-green-800'},
        {'bg': 'bg-purple-100', 'border': 'border-purple-300', 'text': 'text-purple-800'},
        {'bg': 'bg-pink-100', 'border': 'border-pink-300', 'text': 'text-pink-800'},
        {'bg': 'bg-yellow-100', 'border': 'border-yellow-300', 'text': 'text-yellow-800'},
        {'bg': 'bg-teal-100', 'border': 'border-teal-300', 'text': 'text-teal-800'},
        {'bg': 'bg-orange-100', 'border': 'border-orange-300', 'text': 'text-orange-800'},
        {'bg': 'bg-indigo-100', 'border': 'border-indigo-300', 'text': 'text-indigo-800'},
        {'bg': 'bg-rose-100', 'border': 'border-rose-300', 'text': 'text-rose-800'},
        {'bg': 'bg-cyan-100', 'border': 'border-cyan-300', 'text': 'text-cyan-800'},
    ]
    
    group_colors = {}
    for i, group in enumerate(groups):
        color_index = i % len(pastel_colors)
        group_colors[group.id] = pastel_colors[color_index]
    
    for item in schedule_data:
        color = group_colors.get(item['group'].id, pastel_colors[0])
        item['bg_color'] = color['bg']
        item['border_color'] = color['border']
        item['text_color'] = color['text']
    
    context = {
        'days': days,
        'time_slots': time_slots,
        'schedule_data': schedule_data,
        'role': role,
    }
    
    return render(request, 'dashboard/weekly_schedule.html', context)


@login_required
def profile(request):
    """Личный кабинет пользователя"""
    role = get_user_role(request.user)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'change_password':
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if not request.user.check_password(old_password):
                messages.error(request, 'Неверный текущий пароль')
            elif new_password != confirm_password:
                messages.error(request, 'Новые пароли не совпадают')
            elif len(new_password) < 6:
                messages.error(request, 'Пароль должен быть не менее 6 символов')
            else:
                request.user.set_password(new_password)
                request.user.save()
                messages.success(request, 'Пароль успешно изменён!')
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
        
        elif action == 'change_phone':
            phone = request.POST.get('phone')
            profile_obj, created = Profile.objects.get_or_create(user=request.user)
            profile_obj.phone = phone
            profile_obj.save()
            messages.success(request, 'Телефон обновлён!')
        
        elif action == 'change_photo':
            photo = request.FILES.get('photo')
            if photo:
                profile_obj, created = Profile.objects.get_or_create(user=request.user)
                profile_obj.photo = photo
                profile_obj.save()
                messages.success(request, 'Фото обновлено!')
            else:
                messages.error(request, 'Выберите файл')
    
    try:
        user_profile = request.user.profile
    except:
        user_profile = None
    
    context = {
        'role': role,
        'user_profile': user_profile,
    }
    
    return render(request, 'dashboard/profile.html', context)


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


@login_required
def export_excel(request):
    """Экспорт данных в Excel для админа"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'accountant', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(vertical='center')
    
    # Лист 1: Ученики и долги
    ws1 = wb.active
    ws1.title = 'Ученики и долги'
    
    headers1 = ['Имя', 'Телефон', 'Группы (цены)', 'Общий долг (₸)', 'Не оплачено циклов']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    students = Student.objects.all()
    for row, student in enumerate(students, 2):
        enrollments = Enrollment.objects.filter(student=student).select_related('group')
        
        groups_info = []
        for enrollment in enrollments:
            group_price = float(enrollment.group.price)
            groups_info.append(f"{enrollment.group.name} ({group_price}₸)")
        groups_str = ', '.join(groups_info) if groups_info else '—'
        
        total_debt = 0
        cycles_unpaid = 0
        
        for enrollment in enrollments:
            unpaid_payments = Payment.objects.filter(
                student=student,
                group=enrollment.group,
                is_paid=False
            )
            total_debt += sum([float(p.amount) - float(p.paid_amount) for p in unpaid_payments])
            cycles_unpaid += unpaid_payments.count()
        
        data = [
            student.name,
            student.phone or '',
            groups_str,
            total_debt if total_debt > 0 else 0,
            cycles_unpaid,
        ]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
            if col == 4 and total_debt > 0:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                cell.font = Font(color='DC2626', bold=True)
            elif col == 5 and cycles_unpaid > 0:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                cell.font = Font(color='DC2626')
    
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 25
    
    # Лист 2: Долги по циклам
    ws2 = wb.create_sheet('Долги по циклам')
    
    headers2 = ['Ученик', 'Телефон', 'Телефон родителя', 'Группа', 'Цена группы', 'Цикл', 'Период', 'Долг', 'Статус']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    unpaid = Payment.objects.filter(is_paid=False).select_related('student', 'group').order_by('student__name', 'cycle_number')
    for row, payment in enumerate(unpaid, 2):
        period = f"{payment.start_date.strftime('%d.%m')} — {payment.end_date.strftime('%d.%m')}" if payment.start_date and payment.end_date else '—'
        data = [
            payment.student.name,
            payment.student.phone or '—',
            payment.student.parent_phone or '—',
            payment.group.name,
            float(payment.group.price),
            f"Цикл {payment.cycle_number}",
            period,
            float(payment.amount) - float(payment.paid_amount),
            '❌ Не оплачено',
        ]
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
            if col == 7:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                cell.font = Font(color='DC2626')
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 25
    ws2.column_dimensions['H'].width = 15
    ws2.column_dimensions['I'].width = 20
    
    # Лист 3: Группы
    ws3 = wb.create_sheet('Группы')
    
    headers3 = ['Название', 'Учитель', 'Расписание', 'Цена', 'Учеников', 'Активна']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    groups = Group.objects.all()
    for row, group in enumerate(groups, 2):
        data = [
            group.name,
            group.teacher.username if group.teacher else '—',
            group.schedule or '',
            float(group.price),
            group.enrollments.count(),
            'Да' if group.is_active else 'Нет',
        ]
        for col, value in enumerate(data, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 10
    
    # Лист 4: Посещаемость
    ws4 = wb.create_sheet('Посещаемость')
    
    headers4 = ['Дата', 'Группа', 'Ученик', 'Статус']
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    attendances = Attendance.objects.select_related('lesson', 'lesson__group', 'student').order_by('-lesson__date')
    for row, attendance in enumerate(attendances, 2):
        status_map = {
            'present': 'Присутствовал',
            'absent': 'Отсутствовал',
            'late': 'Опоздал',
        }
        data = [
            attendance.lesson.date.strftime('%d.%m.%Y'),
            attendance.lesson.group.name,
            attendance.student.name,
            status_map.get(attendance.status, attendance.status),
        ]
        for col, value in enumerate(data, 1):
            cell = ws4.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
            if col == 4:
                if attendance.status == 'present':
                    cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                elif attendance.status == 'absent':
                    cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['D'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="english_center_report.xlsx"'
    
    wb.save(response)
    return response


@login_required
def payment_management(request):
    """Страница контроля оплат"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'accountant', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    today = timezone.localdate()
    view_mode = request.GET.get('view', 'periods')
    search_query = request.GET.get('search', '')
    
    if view_mode == 'students':
        # РЕЖИМ ПО УЧЕНИКАМ
        students = Student.objects.all()
        
        if search_query:
            students = students.filter(name__icontains=search_query)
        
        students_data = []
        for student in students:
            unpaid_payments = Payment.objects.filter(student=student, is_paid=False)
            total_debt = sum([float(p.amount) - float(p.paid_amount) for p in unpaid_payments])
            
            if total_debt > 0:
                groups_list = []
                for payment in unpaid_payments:
                    if payment.group.name not in groups_list:
                        groups_list.append(payment.group.name)
                
                students_data.append({
                    'student': student,
                    'total_debt': total_debt,
                    'groups': ', '.join(groups_list),
                    'unpaid_count': unpaid_payments.count(),
                    'student_phone': student.phone,
                    'parent_phone': student.parent_phone,
                })
        
        students_data.sort(key=lambda x: x['total_debt'], reverse=True)
        
        context = {
            'view_mode': view_mode,
            'students_data': students_data,
            'role': role,
            'search_query': search_query,
        }
    else:
        # РЕЖИМ ПО ПЕРИОДАМ
        payment_data = []
        
        all_payments = Payment.objects.select_related('student', 'group').order_by('student__name', 'cycle_number')
        
        if search_query:
            all_payments = all_payments.filter(student__name__icontains=search_query)
        
        for payment in all_payments:
            student = payment.student
            group = payment.group
            
            if payment.is_paid:
                status = 'paid'
            elif payment.is_partial:
                status = 'partial'
            elif payment.end_date and today > payment.end_date:
                status = 'overdue'
            elif payment.start_date and payment.start_date <= today <= payment.end_date:
                status = 'current'
            else:
                status = 'upcoming'
            
            payment_data.append({
                'payment': payment,
                'student': student,
                'group': group,
                'cycle_number': payment.cycle_number,
                'start_date': payment.start_date,
                'end_date': payment.end_date,
                'status': status,
                'remaining': float(payment.amount) - float(payment.paid_amount),
                'student_phone': student.phone,
                'parent_name': student.parent_name,
                'parent_phone': student.parent_phone,
            })
        
        status_order = {'overdue': 0, 'current': 1, 'partial': 2, 'upcoming': 3, 'paid': 4}
        payment_data.sort(key=lambda x: (status_order.get(x['status'], 5), x['end_date']))
        
        context = {
            'view_mode': view_mode,
            'payment_data': payment_data,
            'role': role,
            'search_query': search_query,
        }
    
    return render(request, 'dashboard/payment_management.html', context)


@login_required
def toggle_payment_management(request, payment_id):
    """Отметить полную оплату"""
    if request.method == 'POST':
        role = get_user_role(request.user)
        
        if role not in ['admin', 'accountant', 'developer']:
            return JsonResponse({'success': False, 'error': 'Нет доступа'})
        
        payment = get_object_or_404(Payment, id=payment_id)
        
        if payment.is_paid:
            # Отменяем оплату
            payment.is_paid = False
            payment.is_partial = False
            payment.paid_amount = 0
            payment.paid_at = None
            payment.marked_by = None
        else:
            # Полная оплата
            payment.is_paid = True
            payment.is_partial = False
            payment.paid_amount = payment.amount
            payment.paid_at = timezone.now()
            payment.marked_by = request.user
        
        payment.save()
        return JsonResponse({'success': True, 'is_paid': payment.is_paid})
    
    return JsonResponse({'success': False})


@login_required
def partial_payment(request, payment_id):
    """Частичная оплата"""
    if request.method == 'POST':
        role = get_user_role(request.user)
        
        if role not in ['admin', 'accountant', 'developer']:
            return JsonResponse({'success': False, 'error': 'Нет доступа'})
        
        payment = get_object_or_404(Payment, id=payment_id)
        amount = float(request.POST.get('amount', 0))
        
        if amount > 0:
            from decimal import Decimal
            payment.paid_amount = Decimal(str(payment.paid_amount)) + Decimal(str(amount))
            payment.is_partial = True
            payment.marked_by = request.user
            
            if payment.paid_amount >= payment.amount:
                payment.is_paid = True
                payment.is_partial = False
                payment.paid_amount = payment.amount
                payment.paid_at = timezone.now()
            
            payment.save()
            return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

@login_required
def accountant_stats(request):
    """Статистика для бухгалтера"""
    role = get_user_role(request.user)
    
    if role not in ['accountant', 'developer', 'admin']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    today = timezone.localdate()
    
    # Период
    period = request.GET.get('period', 'month')
    group_filter = request.GET.get('group', 'all')
    
    # Все платежи с положительной суммой (включая частичные)
    payments = Payment.objects.filter(paid_amount__gt=0)
    
    if group_filter != 'all':
        payments = payments.filter(group_id=group_filter)
    
    # Общий заработок (все оплаты)
    total_income = sum([float(p.paid_amount) for p in payments])
    
    # По группам
    groups_stats = []
    all_groups = Group.objects.all()
    for group in all_groups:
        group_payments = payments.filter(group=group)
        group_income = sum([float(p.paid_amount) for p in group_payments])
        if group_income > 0:
            groups_stats.append({
                'group': group,
                'income': group_income,
                'count': group_payments.count(),
            })
    
    groups_stats.sort(key=lambda x: x['income'], reverse=True)
    
    # График по месяцам за всё время
    monthly_income = []
    for i in range(11, -1, -1):
        month_date = today.replace(day=1) - timedelta(days=i*30)
        month_start = month_date.replace(day=1)
        if month_start.month == 12:
            month_end = month_start.replace(day=31)
        else:
            next_month = month_start.replace(month=month_start.month + 1, day=1)
            month_end = next_month - timedelta(days=1)
        
        # Оплаты с датой оплаты в этом месяце
        paid_with_date = Payment.objects.filter(
            paid_amount__gt=0,
            paid_at__isnull=False,
            paid_at__date__gte=month_start,
            paid_at__date__lte=month_end
        )
        
        # Частичные оплаты без даты — по дате конца периода
        partial_without_date = Payment.objects.filter(
            paid_amount__gt=0,
            paid_at__isnull=True,
            end_date__gte=month_start,
            end_date__lte=month_end
        )
        
        month_income = sum([float(p.paid_amount) for p in paid_with_date]) + \
                       sum([float(p.paid_amount) for p in partial_without_date])
        
        monthly_income.append({
            'month': month_start.strftime('%B %Y'),
            'income': month_income,
        })
    
    context = {
        'role': role,
        'total_income': total_income,
        'groups_stats': groups_stats,
        'monthly_income': monthly_income,
        'period': period,
        'start_date': today.replace(day=1),
        'end_date': today,
        'group_filter': group_filter,
        'all_groups': all_groups,
    }
    
    return render(request, 'dashboard/accountant_stats.html', context)


@login_required
def export_income_excel(request):
    """Экспорт статистики доходов в Excel"""
    role = get_user_role(request.user)
    
    if role not in ['accountant', 'developer', 'admin']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    today = timezone.localdate()
    period = request.GET.get('period', 'month')
    group_filter = request.GET.get('group', 'all')
    
    if period == 'month':
        start_date = today.replace(day=1)
    elif period == '3months':
        start_date = today - timedelta(days=90)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today.replace(day=1)
    
    end_date = today
    
    payments = Payment.objects.filter(
        is_paid=True,
        paid_at__date__gte=start_date,
        paid_at__date__lte=end_date
    )
    
    if group_filter != 'all':
        payments = payments.filter(group_id=group_filter)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Доходы'
    
    headers = ['Ученик', 'Группа', 'Сумма', 'Дата оплаты', 'Кто отметил']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    for row, payment in enumerate(payments, 2):
        data = [
            payment.student.name,
            payment.group.name,
            float(payment.paid_amount),
            payment.paid_at.strftime('%d.%m.%Y') if payment.paid_at else '—',
            payment.marked_by.username if payment.marked_by else '—',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="income_report.xlsx"'
    wb.save(response)
    return response


@login_required
def add_group(request):
    """Создание группы"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        group_type = request.POST.get('group_type', 'group')
        teacher_id = request.POST.get('teacher')
        schedule = request.POST.get('schedule', '')
        price = request.POST.get('price', '0')
        
        if name and teacher_id:
            group = Group.objects.create(
                name=name,
                group_type=group_type,
                teacher_id=teacher_id,
                schedule=schedule,
                price=price,
                is_active=True,
            )
            messages.success(request, f'Группа "{name}" создана!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Заполните название и выберите учителя')
    
    teachers = User.objects.filter(profile__role='teacher')
    
    context = {
        'teachers': teachers,
    }
    
    return render(request, 'dashboard/add_group.html', context)


@login_required
def remove_student_from_group(request, student_id, group_id):
    """Удалить ученика из группы"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    student = get_object_or_404(Student, id=student_id)
    group = get_object_or_404(Group, id=group_id)
    
    Enrollment.objects.filter(student=student, group=group).delete()
    messages.success(request, f'{student.name} удалён из группы {group.name}')
    
    return redirect('students_list')


@login_required
def delete_student(request, student_id):
    """Удалить ученика из базы"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    student = get_object_or_404(Student, id=student_id)
    name = student.name
    student.delete()
    messages.success(request, f'Ученик {name} удалён из базы')
    
    return redirect('students_list')


@login_required
def delete_group(request, group_id):
    """Удалить группу"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    group = get_object_or_404(Group, id=group_id)
    name = group.name
    group.delete()
    messages.success(request, f'Группа "{name}" удалена')
    
    return redirect('dashboard')


@login_required
def add_existing_student_to_group(request, student_id):
    """Добавить существующего ученика в группу"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        group_ids = request.POST.getlist('groups')
        
        for group_id in group_ids:
            group = Group.objects.get(id=group_id)
            Enrollment.objects.get_or_create(student=student, group=group)
        
        messages.success(request, f'{student.name} добавлен в выбранные группы')
        return redirect('students_list')
    
    groups = Group.objects.filter(is_active=True)
    current_groups = Enrollment.objects.filter(student=student).values_list('group_id', flat=True)
    available_groups = groups.exclude(id__in=current_groups)
    
    context = {
        'student': student,
        'available_groups': available_groups,
    }
    
    return render(request, 'dashboard/add_to_group.html', context)

@login_required
def toggle_book_status(request, enrollment_id=None):
    """Переключить статус книги у ученика"""
    if request.method == 'POST':
        role = get_user_role(request.user)
        
        if role not in ['admin', 'teacher', 'developer']:
            return JsonResponse({'success': False, 'error': 'Нет доступа'})
        
        action = request.POST.get('action')
        
        # Если переданы несколько учеников
        enrollment_ids = request.POST.get('enrollment_ids', '')
        if enrollment_ids:
            ids = enrollment_ids.split(',')
            for eid in ids:
                try:
                    enrollment = Enrollment.objects.get(id=int(eid))
                    if action == 'need_book':
                        enrollment.book_needed = True
                        enrollment.has_book = False
                    elif action == 'has_book':
                        enrollment.has_book = True
                        enrollment.book_needed = False
                    enrollment.save()
                except:
                    pass
            return JsonResponse({'success': True})
        
        # Для одного ученика
        if enrollment_id:
            enrollment = get_object_or_404(Enrollment, id=enrollment_id)
            if action == 'need_book':
                enrollment.book_needed = True
                enrollment.has_book = False
            elif action == 'has_book':
                enrollment.has_book = True
                enrollment.book_needed = False
            enrollment.save()
            return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})


@login_required
def books_status(request):
    """Статус книг для админа и разработчика"""
    role = get_user_role(request.user)
    
    if role not in ['admin', 'developer']:
        messages.error(request, 'У вас нет доступа')
        return redirect('dashboard')
    
    view_mode = request.GET.get('view', 'books')
    search_query = request.GET.get('search', '')
    
    if view_mode == 'students':
        # Вкладка по ученикам
        enrollments = Enrollment.objects.filter(book_needed=True).select_related('student', 'group', 'group__book')
        
        if search_query:
            enrollments = enrollments.filter(
                student__name__icontains=search_query
            ) | enrollments.filter(
                group__name__icontains=search_query
            ) | enrollments.filter(
                group__book__title__icontains=search_query
            )
        
        context = {
            'view_mode': view_mode,
            'enrollments': enrollments,
            'search_query': search_query,
            'role': role,
        }
    else:
        # Вкладка по книгам
        books = Book.objects.all()
        
        book_stats = []
        for book in books:
            groups_using = Group.objects.filter(book=book)
            students_with_book = Enrollment.objects.filter(has_book=True, group__book=book).count()
            students_need_book = Enrollment.objects.filter(book_needed=True, group__book=book).count()
            
            book_stats.append({
                'book': book,
                'groups_using': groups_using.count(),
                'students_with_book': students_with_book,
                'students_need_book': students_need_book,
                'missing': students_need_book - book.quantity,
            })
        
        context = {
            'view_mode': view_mode,
            'book_stats': book_stats,
            'search_query': search_query,
            'role': role,
        }
    
    return render(request, 'dashboard/books_status.html', context)